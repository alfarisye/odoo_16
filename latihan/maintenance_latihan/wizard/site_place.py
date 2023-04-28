import io
import base64
from openpyxl import Workbook
import openpyxl
from odoo import models, fields, api
from odoo.tools import pycompat

class SitePlace(models.TransientModel):
    _name = 'site.place'

    category_id = fields.Many2one(
        comodel_name='maintenance.equipment.category', string='Equipment Categories')
    category_iid = fields.Many2one(
        comodel_name='category', string='Category')
    brand_id = fields.Many2one(
        comodel_name='brand', string='Brand')
    generated_excel_file = fields.Binary(string='Generated Excel File', readonly=True)

    @api.model
    def create_site_place(self):
        site_place = self.env['site.place'].create({
            'name': self.name,
            'address': self.address,
            'city': self.city,
        })
        return {
            'name': 'Site Place Created',
            'view_mode': 'form',
            'res_model': 'site.place',
            'res_id': site_place.id,
            'type': 'ir.actions.act_window',
            'target': 'new',
        }

    def action_generate(self):
        self.ensure_one()
        
        workbook = Workbook()

        worksheet = workbook.active

        headers = [
            "Equipment Categories",
            "Order By",
            "Nama",
        ]
    
        # Write headers
        for i, header in enumerate(headers):
            worksheet.cell(row=1, column=i+1, value=header)

        worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        worksheet['A1'] = "Equipment Categories"
        worksheet['A1'].fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        worksheet['A1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)
        worksheet['B1'] = "Order By"
        worksheet['B1'].fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        worksheet['B1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Write headers for Category and Brand
        worksheet['B2'] = "Category"
        worksheet['B2'].fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        worksheet['B2'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet['C2'] = "Brand"
        worksheet['C2'].fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        worksheet['C2'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=4, end_row=2, end_column=4)
        worksheet['D1'] = "Nama"
        worksheet['D1'].fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        worksheet['D1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)



        domain = [('category_id.id', '=', self.category_id.id)]

        if self.category_iid:
            domain += [('category_iid.id', '=', self.category_iid.id)]
        
        if self.brand_id:
            domain += [('brand_id.id', '=', self.brand_id.id)]
        equipment_ids = self.env['maintenance.equipment'].sudo().search(domain) 

        row = 3
        for equipment in equipment_ids:
            worksheet.cell(row=row, column=1, value=equipment.category_id.name or '')
            worksheet.cell(row=row, column=2, value=equipment.category_iid.name or '')
            worksheet.cell(row=row, column=3, value=equipment.brand_id.name or '')
            worksheet.cell(row=row, column=4, value=equipment.name or '')
            row += 1

        output = io.BytesIO()
        workbook.save(output)

        self.generated_excel_file = base64.b64encode(output.getvalue())

        filename = "generated_file.xlsx"
        url = "/web/content/site.place/%s/generated_excel_file/%s?download=true" % (self.id, pycompat.to_text(filename))
        workbook.save(output)
        return {
            'type': 'ir.actions.act_url',
            'url': url,
            'target': 'self',
        }
